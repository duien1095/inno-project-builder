#!/usr/bin/env python3
"""Module sinh Excel từ JSON template (cấu trúc danh mục dữ liệu).

Có thể dùng như module hoặc CLI:
    python -m src.excel_generator tem_ex_qhc.json [output.xlsx]
"""

import json
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import (
    Font, Alignment, PatternFill, Border, Side,
)
from openpyxl.utils import get_column_letter

# ═══════════════════════════ Màu sắc ═══════════════════════════

NAVY = "1F4E78"
LIGHT_BLUE = "D9E1F2"
WHITE = "FFFFFF"
GRAY_LINE = "BFBFBF"
DARK_GRAY = "333333"

# ═══════════════════════════ Styles ═══════════════════════════

# Title block (khối tiêu đề đầu sheet) — nền xanh nhạt, chữ tối
TITLE_BLOCK_FONT = Font(name="Arial", size=12, bold=True, color=DARK_GRAY)
TITLE_BLOCK_FILL = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
TITLE_BLOCK_ALIGNMENT = Alignment(horizontal="left", vertical="center", wrap_text=True)

# Header cột — nền navy, chữ trắng
HEADER_FONT = Font(name="Arial", size=10, bold=True, color=WHITE)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
HEADER_FILL = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")

# Cấp 1 — nền xanh nhạt, chữ đậm
LEVEL1_FONT = Font(name="Arial", size=10, bold=True, color=DARK_GRAY)
LEVEL1_FILL = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
LEVEL1_ALIGNMENT = Alignment(horizontal="left", vertical="center", wrap_text=True)

# Cấp 2 — không nền, chữ thường, thụt lề
LEVEL2_FONT = Font(name="Arial", size=10, color=DARK_GRAY)
LEVEL2_FILL = PatternFill()
LEVEL2_ALIGNMENT = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=2)

CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

THIN_SIDE = Side(style="thin", color=GRAY_LINE)
MEDIUM_SIDE = Side(style="medium", color=NAVY)

COLUMN_WIDTHS = {
    "STT": 8,
    "Hạng mục": 42,
    "Tên dữ liệu": 42,
    "Tình trạng": 18,
    "Ngày nhận": 15,
    "Nguồn cung cấp": 28,
    "Ghi chú": 22,
    "Người cập nhật": 22,
    "Nội dung cập nhật": 40,
    "Ngày gửi": 15,
}
DEFAULT_WIDTH = 18

# ═══════════════════════════ Mapping với template_engine ═══════════════════════════

EXCEL_TEMPLATE_NAMES = {"QHC", "QHPK", "QHCT"}

EXCEL_JSON_MAP = {
    "QHC": "tem_ex_qhc.json",
    "QHPK": "tem_ex_qhpk.json",
    "QHCT": "tem_ex_qhct.json",
}

EXCEL_OUTPUT_FILENAME = "BangTinhChiTieu.xlsx"
TARGET_SUBDIR = "02.PROCESS"


# ═══════════════════════════ Hàm helper ═══════════════════════════


def _get_col_letter(col_index: int) -> str:
    return get_column_letter(col_index)


def _find_col_index(columns: list, target: str) -> int | None:
    """Tìm index (1-based) của cột có tên target trong danh sách columns."""
    for idx, col_name in enumerate(columns, 1):
        if col_name == target:
            return idx
    return None


def _apply_border(ws, min_row, max_row, min_col, max_col):
    """Apply border. Viền ngoài medium navy, viền trong thin gray."""
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = Border(
                left=THIN_SIDE, right=THIN_SIDE,
                top=THIN_SIDE, bottom=THIN_SIDE,
            )

    # Viền ngoài dày hơn
    for col in range(min_col, max_col + 1):
        c = ws.cell(row=min_row, column=col)
        c.border = Border(
            left=c.border.left, right=c.border.right,
            top=MEDIUM_SIDE, bottom=c.border.bottom,
        )
        c = ws.cell(row=max_row, column=col)
        c.border = Border(
            left=c.border.left, right=c.border.right,
            top=c.border.top, bottom=MEDIUM_SIDE,
        )
    for row in range(min_row, max_row + 1):
        c = ws.cell(row=row, column=min_col)
        c.border = Border(
            left=MEDIUM_SIDE, right=c.border.right,
            top=c.border.top, bottom=c.border.bottom,
        )
        c = ws.cell(row=row, column=max_col)
        c.border = Border(
            left=c.border.left, right=MEDIUM_SIDE,
            top=c.border.top, bottom=c.border.bottom,
        )


# ═══════════════════════════ Sheet danh mục ═══════════════════════════

def _build_danh_muc_sheet(ws, sheet_data: dict, project_name: str):
    """Ghi sheet DANH_MUC_DU_LIEU."""
    columns = sheet_data.get("columns", [])
    rows_data = sheet_data.get("rows", [])
    title_block = sheet_data.get("title_block", {})
    col_count = len(columns)

    # Tìm cột "Hạng mục" và "Tên dữ liệu" từ columns
    hang_muc_col = _find_col_index(columns, "Hạng mục")
    ten_du_lieu_col = _find_col_index(columns, "Tên dữ liệu")
    if hang_muc_col is None:
        hang_muc_col = 2  # fallback
    if ten_du_lieu_col is None:
        ten_du_lieu_col = 3  # fallback

    # ── Title block — nền xanh nhạt (#D9E1F2), chữ tối ──
    current_row = 1
    if title_block:
        title_lines = title_block.get("lines", [])

        for i, line in enumerate(title_lines):
            r = current_row + i
            cell = ws.cell(row=r, column=1, value=line)
            cell.font = TITLE_BLOCK_FONT
            cell.fill = TITLE_BLOCK_FILL
            cell.alignment = TITLE_BLOCK_ALIGNMENT

            if col_count > 1:
                sc = _get_col_letter(1)
                ec = _get_col_letter(col_count)
                ws.merge_cells(f"{sc}{r}:{ec}{r}")

        current_row += len(title_lines)

    # ── Header (ngay sát dưới title block, không space) ──
    header_row = current_row
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=current_row, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT

    current_row += 1

    # ── Dữ liệu ──
    for row_data in rows_data:
        stt = str(row_data.get("stt", ""))
        name = row_data.get("name", "")
        level = row_data.get("level", 1)

        # STT (cột 1)
        cell_stt = ws.cell(row=current_row, column=1, value=stt)
        cell_stt.font = LEVEL1_FONT if level == 1 else LEVEL2_FONT
        cell_stt.fill = LEVEL1_FILL if level == 1 else LEVEL2_FILL
        cell_stt.alignment = CENTER_ALIGNMENT

        # Cột "Hạng mục" — ghi name vào đây (không phải Tên dữ liệu)
        cell_hm = ws.cell(row=current_row, column=hang_muc_col, value=name)
        if level == 1:
            cell_hm.font = LEVEL1_FONT
            cell_hm.fill = LEVEL1_FILL
            cell_hm.alignment = LEVEL1_ALIGNMENT
        else:
            cell_hm.font = LEVEL2_FONT
            cell_hm.fill = LEVEL2_FILL
            cell_hm.alignment = LEVEL2_ALIGNMENT

        # Cột "Tên dữ liệu" — để trống (dành cho người dùng tự điền)
        # Chỉ apply styling

        # Các cột còn lại
        for col_idx in range(1, col_count + 1):
            if col_idx in (1, hang_muc_col):
                continue
            c = ws.cell(row=current_row, column=col_idx)
            c.font = LEVEL1_FONT if level == 1 else LEVEL2_FONT
            c.fill = LEVEL1_FILL if level == 1 else LEVEL2_FILL
            c.alignment = (
                CENTER_ALIGNMENT if col_idx == 1 else LEVEL1_ALIGNMENT
            )

        current_row += 1

    data_end_row = current_row - 1

    # ── Border ──
    _apply_border(ws, header_row, data_end_row, 1, col_count)

    if title_block:
        title_end_row = header_row - 1
        _apply_border(ws, 1, title_end_row, 1, col_count)

    # ── Column widths ──
    for col_idx, col_name in enumerate(columns, 1):
        width = COLUMN_WIDTHS.get(col_name, DEFAULT_WIDTH)
        ws.column_dimensions[_get_col_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 30
    for r in range(2, current_row):
        ws.row_dimensions[r].height = 20


# ═══════════════════════════ Sheet nhật ký ═══════════════════════════

def _build_nhat_ky_sheet(ws, sheet_data: dict, project_name: str):
    """Ghi sheet NHAT_KY_CAP_NHAT."""
    columns = sheet_data.get("columns", [])
    title_block = sheet_data.get("title_block", {})
    empty_rows = sheet_data.get("empty_rows_for_entry", 0)
    col_count = len(columns)

    current_row = 1
    if title_block:
        title_lines = title_block.get("lines", [])

        for i, line in enumerate(title_lines):
            r = current_row + i
            cell = ws.cell(row=r, column=1, value=line)
            cell.font = TITLE_BLOCK_FONT
            cell.fill = TITLE_BLOCK_FILL
            cell.alignment = TITLE_BLOCK_ALIGNMENT

            if col_count > 1:
                sc = _get_col_letter(1)
                ec = _get_col_letter(col_count)
                ws.merge_cells(f"{sc}{r}:{ec}{r}")

        current_row += len(title_lines)

    # Header — ngay sát dưới title block
    header_row = current_row
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=current_row, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT

    current_row += 1

    # Dòng trống
    for _ in range(empty_rows):
        for col_idx in range(1, col_count + 1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.font = Font(name="Arial", size=10, color=DARK_GRAY)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        current_row += 1

    empty_end = current_row - 1

    # Border
    _apply_border(ws, header_row, empty_end, 1, col_count)
    if title_block:
        title_end_row = header_row - 1
        _apply_border(ws, 1, title_end_row, 1, col_count)

    # Column widths
    for col_idx, col_name in enumerate(columns, 1):
        width = COLUMN_WIDTHS.get(col_name, DEFAULT_WIDTH)
        ws.column_dimensions[_get_col_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 30


# ═══════════════════════════ API chính ═══════════════════════════

def generate_excel_from_template(
    json_path: Path,
    output_path: Path,
    project_name: str = "",
) -> Path:
    """Đọc file JSON và sinh file Excel tương ứng.

    API tương thích với template_engine.

    Args:
        json_path: Đường dẫn file JSON template.
        output_path: Đường dẫn file Excel đầu ra.
        project_name: Tên dự án (hiển thị trong file).

    Returns:
        Đường dẫn file Excel đã tạo.
    """
    if not json_path.exists():
        raise FileNotFoundError(f"Không tìm thấy file: {json_path}")

    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    excel_templates = data.get("excel_templates", [])
    if not excel_templates:
        raise ValueError("File JSON không có 'excel_templates'.")

    tmpl = excel_templates[0]
    template_name = tmpl.get("template_name", "UNKNOWN")
    sheets_data = tmpl.get("sheets", [])

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    for sheet_data in sheets_data:
        sheet_name = sheet_data.get("sheet_name", "Sheet")

        if sheet_name == "DANH_MUC_DU_LIEU":
            if wb.active:
                ws = wb.active
                ws.title = sheet_name
            else:
                ws = wb.create_sheet(title=sheet_name)
            _build_danh_muc_sheet(ws, sheet_data, project_name or template_name)
        else:
            ws = wb.create_sheet(title=sheet_name)
            _build_nhat_ky_sheet(ws, sheet_data, project_name or template_name)

    wb.save(str(output_path))
    return output_path


def _find_excel_json(template_name: str) -> Path | None:
    """Tìm file JSON tương ứng với template."""
    json_filename = EXCEL_JSON_MAP.get(template_name)
    if not json_filename:
        return None

    candidates = [
        Path.cwd() / json_filename,
        Path(__file__).resolve().parent.parent / json_filename,
    ]

    for candidate in candidates:
        if candidate.exists() and candidate.is_file():
            return candidate

    return None


def _find_target_folder(project_folder: Path, folder_paths: list[str]) -> Path | None:
    """Tìm thư mục PROCESS trong folder_paths."""
    if not folder_paths:
        return None

    for folder_path in folder_paths:
        name_upper = folder_path.upper()
        if "PROCESS" in name_upper or "INPUT" in name_upper:
            target = project_folder / folder_path
            if target.exists() and target.is_dir():
                return target

    first = project_folder / folder_paths[0]
    if first.exists() and first.is_dir():
        return first

    return None


def generate_excel_for_project(
    project_folder: Path,
    template_name: str,
    folder_paths: list[str],
    created_files: list[Path] | None = None,
) -> Path | None:
    """Sinh file Excel từ JSON template và lưu vào thư mục dự án.

    API tương thích với template_engine.
    Chỉ sinh cho QHC, QHPK, QHCT. Bỏ qua CONCEPT_*.
    """
    if template_name not in EXCEL_TEMPLATE_NAMES:
        return None

    if created_files is None:
        created_files = []

    json_path = _find_excel_json(template_name)
    if json_path is None:
        return None

    # Tìm thư mục PROCESS
    target_dir = _find_target_folder(project_folder, folder_paths)
    if target_dir is None:
        target_dir = project_folder / TARGET_SUBDIR
        target_dir.mkdir(parents=True, exist_ok=True)

    try:
        with open(json_path, "r", encoding="utf-8") as f:
            data = json.load(f)

        excel_templates = data.get("excel_templates", [])
        file_name = EXCEL_OUTPUT_FILENAME
        if excel_templates and excel_templates[0].get("file_name"):
            file_name = excel_templates[0]["file_name"]

        output_path = target_dir / file_name

        generate_excel_from_template(
            json_path=json_path,
            output_path=output_path,
            project_name=project_folder.name,
        )
        if output_path.exists():
            created_files.append(output_path)
            return output_path
    except Exception:
        pass

    return None


# ═══════════════════════════ CLI ═══════════════════════════

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    json_path = Path(sys.argv[1])
    output_path = Path(sys.argv[2]) if len(sys.argv) > 2 else None

    try:
        result = generate_excel_from_template(
            json_path,
            output_path or Path("output.xlsx"),
            project_name="",
        )
        wb = load_workbook(result)
        print(f"Excel da tao: {result}")
        print(f"   Kich thuoc: {result.stat().st_size:,} bytes")
        print(f"   Sheets: {', '.join(wb.sheetnames)}")
    except Exception as e:
        print(f"Loi: {e}")
        sys.exit(1)
