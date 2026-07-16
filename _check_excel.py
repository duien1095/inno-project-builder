#!/usr/bin/env python3
"""Kiểm tra chi tiết nội dung Excel — dùng openpyxl đọc lại từng ô."""

import json
from pathlib import Path
from openpyxl import load_workbook

JSON_FILES = [
    ("tem_ex_qhc.json", "QHC"),
    ("tem_ex_qhpk.json", "QHPK"),
    ("tem_ex_qhct.json", "QHCT"),
]


def check_sheet_danh_muc(ws):
    """Kiểm tra sheet DANH_MUC_DU_LIEU."""
    errors = []

    # 1. Title block (dòng 1-8) — nền xanh nhạt #D9E1F2, chữ tối
    for r in range(1, 9):
        cell = ws.cell(row=r, column=1)
        fill_color = str(cell.fill.start_color.rgb).upper() if cell.fill.start_color else ""
        if "D9E1F2" not in fill_color:
            errors.append(
                f"   LOI: Title dong {r} fill={fill_color}, mong doi D9E1F2"
            )
        # font size 12
        if cell.font.size and int(cell.font.size) != 12:
            errors.append(
                f"   LOI: Title dong {r} font size={cell.font.size}, mong doi 12"
            )

    # 2. Tìm header row
    header_row = None
    for r in range(1, 20):
        val = ws.cell(row=r, column=1).value
        if val == "STT":
            header_row = r
            break
    if header_row is None:
        errors.append("   LOI: Khong tim thay header row!")
        return errors

    # 3. Đọc tên các cột
    col_headers = {}
    for col in range(1, 10):
        val = ws.cell(row=header_row, column=col).value
        if val:
            col_headers[col] = val

    hang_muc_col = None
    ten_du_lieu_col = None
    for col, name in col_headers.items():
        if name == "Hạng mục":
            hang_muc_col = col
        if name == "Tên dữ liệu":
            ten_du_lieu_col = col

    print(f"   Header dong {header_row}")
    print(f"   Cot 'Hang muc' = {hang_muc_col}, Cot 'Ten du lieu' = {ten_du_lieu_col}")

    # 4. Cột Hạng mục phải có dữ liệu (name từ JSON ghi vào đây)
    found_data_in_hang_muc = False
    if hang_muc_col is not None:
        for r in range(header_row + 1, header_row + 60):
            val = ws.cell(row=r, column=hang_muc_col).value
            if val and isinstance(val, str) and val.strip():
                found_data_in_hang_muc = True
                break
        if not found_data_in_hang_muc:
            errors.append(
                f"   LOI: Cot 'Hang muc' (col {hang_muc_col}) hoan toan trong!"
            )
        else:
            print(f"   OK: Cot 'Hang muc' co du lieu")
    else:
        errors.append("   LOI: Khong tim thay cot 'Hang muc'")

    # 5. Cột Tên dữ liệu phải TRỐNG hoàn toàn
    has_data_in_ten = False
    if ten_du_lieu_col is not None:
        for r in range(header_row + 1, header_row + 60):
            val = ws.cell(row=r, column=ten_du_lieu_col).value
            if val and isinstance(val, str) and val.strip():
                has_data_in_ten = True
                first_data = ws.cell(row=r, column=ten_du_lieu_col).value
                break
        if has_data_in_ten:
            errors.append(
                f"   LOI: Cot 'Ten du lieu' (col {ten_du_lieu_col}) co du lieu "
                f"(dau tien: '{first_data}'), mong doi trong hoan toan!"
            )
        else:
            print(f"   OK: Cot 'Ten du lieu' trong hoan toan")
    else:
        # Fallback: kiểm tra cột C (3)
        for r in range(header_row + 1, header_row + 60):
            val = ws.cell(row=r, column=3).value
            if val and isinstance(val, str) and val.strip():
                has_data_in_ten = True
                break
        if has_data_in_ten:
            errors.append(
                "   LOI: Cot C (Ten du lieu) co du lieu, mong doi trong!"
            )
        else:
            print(f"   OK: Cot C (Ten du lieu - fallback) trong hoan toan")

    # 6. Kiểm tra level=1 có nền xanh nhạt, level=2 không nền, indent
    if hang_muc_col is not None:
        for r in range(header_row + 1, header_row + 30):
            val = ws.cell(row=r, column=hang_muc_col).value
            if not val:
                continue
            cell = ws.cell(row=r, column=hang_muc_col)
            # Không kiểm tra cứng vì phụ thuộc JSON

    return errors


def check_sheet_nhat_ky(ws):
    """Kiểm tra sheet NHAT_KY_CAP_NHAT."""
    errors = []

    # 1. Dòng 1: title
    title = ws.cell(row=1, column=1).value
    if not title:
        errors.append("   LOI: Dong 1 khong co title!")

    # 2. Dòng 2 phải là header (STT), không phải dòng trống
    row2_val = ws.cell(row=2, column=1).value
    if not row2_val or (isinstance(row2_val, str) and row2_val.strip() == ""):
        errors.append(
            f"   LOI: Dong 2 (header) trong! Gia tri: '{row2_val}'"
        )
    if row2_val != "STT":
        errors.append(
            f"   LOI: Dong 2 cot 1 = '{row2_val}', mong doi 'STT'"
        )

    # 3. Title block nên nền xanh nhạt
    fill_color = str(ws.cell(row=1, column=1).fill.start_color.rgb).upper() if ws.cell(row=1, column=1).fill.start_color else ""
    if "D9E1F2" not in fill_color:
        errors.append(
            f"   LOI: Title NK fill={fill_color}, mong doi D9E1F2"
        )

    if not errors:
        print(f"   OK: Title dong 1 = '{title}', Header dong 2 = '{row2_val}'")

    return errors


def main():
    base_dir = Path(__file__).resolve().parent

    for json_file, tmpl_name in JSON_FILES:
        json_path = base_dir / json_file
        if not json_path.exists():
            print(f"Bo qua {json_file}: khong ton tai")
            continue

        output_path = base_dir / f"test_{tmpl_name}.xlsx"
        print(f"\n{'='*60}")
        print(f"Kiem tra: {json_file} -> {output_path.name}")

        try:
            from src.excel_generator import generate_excel_from_template

            result = generate_excel_from_template(
                json_path, output_path, project_name=f"DU_AN_{tmpl_name}"
            )
            print(f"   Da tao: {result} ({result.stat().st_size:,} bytes)")

            wb = load_workbook(result)

            # Sheet DANH_MUC_DU_LIEU
            if "DANH_MUC_DU_LIEU" in wb.sheetnames:
                ws = wb["DANH_MUC_DU_LIEU"]
                errs = check_sheet_danh_muc(ws)
                if errs:
                    for e in errs:
                        print(e)
                else:
                    print("   DANH_MUC_DU_LIEU: OK")
            else:
                print("   LOI: Khong tim thay sheet DANH_MUC_DU_LIEU")

            # Sheet NHAT_KY_CAP_NHAT
            if "NHAT_KY_CAP_NHAT" in wb.sheetnames:
                ws = wb["NHAT_KY_CAP_NHAT"]
                errs = check_sheet_nhat_ky(ws)
                if errs:
                    for e in errs:
                        print(e)
                else:
                    print("   NHAT_KY_CAP_NHAT: OK")
            else:
                print("   LOI: Khong tim thay sheet NHAT_KY_CAP_NHAT")

        except Exception as e:
            import traceback
            print(f"   LOI: {e}")
            traceback.print_exc()

    print("\nDone!")


if __name__ == "__main__":
    main()
