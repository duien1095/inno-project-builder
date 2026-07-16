"""Test cho chức năng sinh Excel từ JSON template."""

import json
import tempfile
import unittest
from pathlib import Path

from src.template_engine import create_template_folders
from src.excel_generator import (
    generate_excel_for_project,
    generate_excel_from_template,
    EXCEL_TEMPLATE_NAMES,
    EXCEL_JSON_MAP,
    EXCEL_OUTPUT_FILENAME,
)


class ExcelTemplateTests(unittest.TestCase):
    """Kiểm tra sinh file Excel từ JSON template."""

    def setUp(self) -> None:
        self.temp_dir = tempfile.TemporaryDirectory()
        self.origin_dir = Path.cwd()

        import os
        os.chdir(str(Path(__file__).resolve().parent.parent))

    def tearDown(self) -> None:
        import os
        os.chdir(str(self.origin_dir))
        self.temp_dir.cleanup()

    def _prepare_project(self, name: str) -> Path:
        """Tạo project folder tạm."""
        project = Path(self.temp_dir.name) / name
        return project

    def test_all_json_files_exist(self) -> None:
        """Tất cả file JSON template đều tồn tại."""
        for tmpl_name, fname in EXCEL_JSON_MAP.items():
            path = Path(fname)
            self.assertTrue(path.exists(), f"Thiếu file {fname} cho {tmpl_name}")

    def test_all_json_parseable(self) -> None:
        """Tất cả file JSON có cấu trúc đúng."""
        for tmpl_name, fname in EXCEL_JSON_MAP.items():
            with open(fname, "r", encoding="utf-8") as f:
                data = json.load(f)
            self.assertIn("excel_templates", data)
            self.assertGreater(len(data["excel_templates"]), 0)
            tmpl = data["excel_templates"][0]
            self.assertIn("sheets", tmpl)
            self.assertGreater(len(tmpl["sheets"]), 0)

            for sheet in tmpl["sheets"]:
                self.assertIn("columns", sheet)
                self.assertGreater(len(sheet["columns"]), 0)

    def test_generate_all_templates(self) -> None:
        """Sinh Excel thành công cho cả 3 template."""
        for tmpl_name in ["QHC", "QHPK", "QHCT"]:
            project = self._prepare_project(f"DU_AN_{tmpl_name}")
            (project / "02.PROCESS").mkdir(parents=True, exist_ok=True)

            result = generate_excel_for_project(
                project, tmpl_name, ["02.PROCESS"]
            )
            self.assertIsNotNone(result, f"Không sinh Excel cho {tmpl_name}")
            self.assertTrue(result.exists(), f"File Excel không tồn tại cho {tmpl_name}")
            self.assertGreater(result.stat().st_size, 1000)

    def test_excel_qhc(self) -> None:
        """Sinh Excel cho QHC."""
        project = self._prepare_project("DU_AN_QHC")
        (project / "02.PROCESS").mkdir(parents=True, exist_ok=True)

        result = generate_excel_for_project(project, "QHC", ["02.PROCESS"])
        self.assertIsNotNone(result)
        self.assertEqual(result.name, "00_DanhMucDuLieu_qhc.xlsx")

    def test_excel_qhpk(self) -> None:
        """Sinh Excel cho QHPK."""
        project = self._prepare_project("DU_AN_QHPK")
        (project / "02.PROCESS").mkdir(parents=True, exist_ok=True)

        result = generate_excel_for_project(project, "QHPK", ["02.PROCESS"])
        self.assertIsNotNone(result)
        self.assertEqual(result.name, "00_DanhMucDuLieu_qhpk.xlsx")

    def test_excel_qhct(self) -> None:
        """Sinh Excel cho QHCT."""
        project = self._prepare_project("DU_AN_QHCT")
        (project / "02.PROCESS").mkdir(parents=True, exist_ok=True)

        result = generate_excel_for_project(project, "QHCT", ["02.PROCESS"])
        self.assertIsNotNone(result)
        self.assertEqual(result.name, "00_DanhMucDuLieu_qhct.xlsx")

    def test_not_generate_for_concept(self) -> None:
        """Không sinh Excel cho CONCEPT."""
        for concept_name in ["CONCEPT_QHPK", "CONCEPT_QHCT"]:
            project = self._prepare_project(f"DU_AN_{concept_name}")
            result = generate_excel_for_project(project, concept_name, ["02.PROCESS"])
            self.assertIsNone(result)

    def test_created_files_list_includes_excel(self) -> None:
        """Kiểm tra created_files có chứa file Excel."""
        project = self._prepare_project("DU_AN_FILES")
        (project / "02.PROCESS").mkdir(parents=True, exist_ok=True)

        created_files: list[Path] = []
        generate_excel_for_project(project, "QHC", ["02.PROCESS"], created_files)

        self.assertEqual(len(created_files), 1)
        self.assertTrue(created_files[0].name.endswith(".xlsx"))

    def test_excel_has_sheets(self) -> None:
        """Kiểm tra Excel có đúng 2 sheet."""
        from openpyxl import load_workbook

        project = self._prepare_project("DU_AN_SHEETS")
        (project / "02.PROCESS").mkdir(parents=True, exist_ok=True)

        generate_excel_for_project(project, "QHC", ["02.PROCESS"])
        xlsx_path = list((project / "02.PROCESS").glob("*.xlsx"))[0]

        wb = load_workbook(xlsx_path)
        self.assertEqual(len(wb.sheetnames), 2)
        self.assertIn("DANH_MUC_DU_LIEU", wb.sheetnames)
        self.assertIn("NHAT_KY_CAP_NHAT", wb.sheetnames)

    def test_excel_header_and_rows(self) -> None:
        """Header và rows được ghi đúng vị trí."""
        from openpyxl import load_workbook

        project = self._prepare_project("DU_AN_ROWS")
        (project / "02.PROCESS").mkdir(parents=True, exist_ok=True)

        generate_excel_for_project(project, "QHC", ["02.PROCESS"])
        xlsx_path = list((project / "02.PROCESS").glob("*.xlsx"))[0]

        wb = load_workbook(xlsx_path)
        ws = wb["DANH_MUC_DU_LIEU"]

        # Title block = 8 dòng, header ở dòng 9
        self.assertEqual(ws.cell(row=9, column=1).value, "STT",
                         "Header STT phai o dong 9")
        self.assertEqual(ws.cell(row=9, column=2).value, "Hạng mục",
                         "Header Hang muc phai o dong 9 cot B")

        # Title block dòng 1
        self.assertEqual(ws.cell(row=1, column=1).value, "DANH MỤC DỮ LIỆU",
                         "Title dong 1 sai")

        # Kiểm tra có dữ liệu trong cột B (Hạng mục)
        has_data = False
        for r in range(10, 50):
            val = ws.cell(row=r, column=2).value
            if val and isinstance(val, str) and val.strip():
                has_data = True
                break
        self.assertTrue(has_data, "Khong co du lieu trong cot Hạng mục!")

    def test_excel_name_in_hang_muc_column(self) -> None:
        """'name' từ JSON phải vào cột 'Hạng mục' (B), cột 'Tên dữ liệu' (C) trống."""
        from openpyxl import load_workbook

        project = self._prepare_project("DU_AN_NAME_COL")
        (project / "02.PROCESS").mkdir(parents=True, exist_ok=True)

        generate_excel_for_project(project, "QHC", ["02.PROCESS"])
        xlsx_path = list((project / "02.PROCESS").glob("*.xlsx"))[0]

        wb = load_workbook(xlsx_path)
        ws = wb["DANH_MUC_DU_LIEU"]

        # Cột B (Hạng mục) phải có dữ liệu
        found_hang_muc = False
        for r in range(10, 50):
            val = ws.cell(row=r, column=2).value
            if val and isinstance(val, str) and val.strip():
                found_hang_muc = True
                break
        self.assertTrue(found_hang_muc, "Cot 'Hang muc' hoan toan trong!")

        # Cột C (Tên dữ liệu) phải trống
        for r in range(10, 50):
            val = ws.cell(row=r, column=3).value
            if val and isinstance(val, str) and val.strip():
                self.fail(f"Cot 'Ten du lieu' (C) dong {r} co du lieu "
                          f"'{val}' trong khi mong doi trong!")

    def test_title_block_light_blue(self) -> None:
        """Title block nền xanh nhạt #D9E1F2, font size 12, chữ tối."""
        from openpyxl import load_workbook

        project = self._prepare_project("DU_AN_TITLE_STYLE")
        (project / "02.PROCESS").mkdir(parents=True, exist_ok=True)

        generate_excel_for_project(project, "QHC", ["02.PROCESS"])
        xlsx_path = list((project / "02.PROCESS").glob("*.xlsx"))[0]

        wb = load_workbook(xlsx_path)
        ws = wb["DANH_MUC_DU_LIEU"]

        for r in range(1, 9):
            cell = ws.cell(row=r, column=1)
            fill = str(cell.fill.start_color.rgb).upper() if cell.fill.start_color else ""
            self.assertIn("D9E1F2", fill,
                          f"Title dong {r} khong phai xanh nhat: {fill}")
            self.assertEqual(cell.font.size, 12,
                             f"Title dong {r} font size = {cell.font.size}, mong doi 12")
            self.assertTrue(cell.font.bold,
                            f"Title dong {r} khong in dam")

    def test_header_navy(self) -> None:
        """Header cột nền navy #1F4E78, chữ trắng."""
        from openpyxl import load_workbook

        project = self._prepare_project("DU_AN_HEADER_STYLE")
        (project / "02.PROCESS").mkdir(parents=True, exist_ok=True)

        generate_excel_for_project(project, "QHC", ["02.PROCESS"])
        xlsx_path = list((project / "02.PROCESS").glob("*.xlsx"))[0]

        wb = load_workbook(xlsx_path)
        ws = wb["DANH_MUC_DU_LIEU"]

        # Header ở dòng 9
        for col in range(1, 8):
            cell = ws.cell(row=9, column=col)
            if cell.value:
                fill = str(cell.fill.start_color.rgb).upper() if cell.fill.start_color else ""
                self.assertIn("1F4E78", fill,
                              f"Header dong 9 cot {col} khong phai navy: {fill}")
                self.assertEqual(cell.font.color.rgb.upper() if cell.font.color else "",
                                 "00FFFFFF" if col == 1 else "00FFFFFF",
                                 f"Header dong 9 cot {col} chu khong trang")

    def test_nhat_ky_no_empty_row(self) -> None:
        """Sheet NHAT_KY_CAP_NHAT không có dòng trống giữa title và header."""
        from openpyxl import load_workbook

        project = self._prepare_project("DU_AN_NK")
        (project / "02.PROCESS").mkdir(parents=True, exist_ok=True)

        generate_excel_for_project(project, "QHC", ["02.PROCESS"])
        xlsx_path = list((project / "02.PROCESS").glob("*.xlsx"))[0]

        wb = load_workbook(xlsx_path)
        ws = wb["NHAT_KY_CAP_NHAT"]

        # Dòng 1: title
        self.assertIsNotNone(ws.cell(row=1, column=1).value,
                             "Dong 1 phai co title")
        # Dòng 2: header
        row2_val = ws.cell(row=2, column=1).value
        self.assertIsNotNone(row2_val,
                             "Dong 2 (header) khong duoc trong")
        self.assertEqual(row2_val, "STT",
                         f"Dong 2 cot 1 = '{row2_val}', mong doi 'STT'")

    def test_integration_create_template_folders_qhc(self) -> None:
        """Tích hợp: create_template_folders tạo thư mục + Excel."""
        project = Path(self.temp_dir.name) / "DU_AN_INTEGRATION"

        created_dirs, created_files = create_template_folders(
            project,
            ["02.PROCESS"],
            template_name="QHC",
        )

        self.assertEqual(len(created_dirs), 1)
        self.assertTrue((project / "02.PROCESS").exists())

        excel_files = [f for f in created_files if f.suffix == ".xlsx"]
        self.assertEqual(len(excel_files), 1)

    def test_integration_with_subfolders(self) -> None:
        """Tích hợp: tạo thư mục đầy đủ + Excel."""
        project = Path(self.temp_dir.name) / "DU_AN_FULL"

        created_dirs, created_files = create_template_folders(
            project,
            ["01.INPUT", "02.PROCESS"],
            template_name="QHC",
        )

        self.assertEqual(len(created_dirs), 2)
        excel_files = [f for f in created_files if f.suffix == ".xlsx"]
        self.assertEqual(len(excel_files), 1)
        self.assertTrue(any(f.suffix == ".pdf" for f in created_files))

    def test_excel_not_generated_for_concept_in_integration(self) -> None:
        """Tích hợp: CONCEPT không sinh Excel."""
        project = Path(self.temp_dir.name) / "DU_AN_CONCEPT"

        created_dirs, created_files = create_template_folders(
            project,
            ["01.INPUT", "02.PROCESS"],
            template_name="CONCEPT_QHPK",
        )

        self.assertEqual(len(created_dirs), 2)
        excel_files = [f for f in created_files if f.suffix == ".xlsx"]
        self.assertEqual(len(excel_files), 0)


if __name__ == "__main__":
    unittest.main()
